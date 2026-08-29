import os
import io
import httpx
from datetime import datetime
from fastapi import HTTPException
import openpyxl

MS_TENANT_ID = os.getenv("MS_TENANT_ID")
MS_CLIENT_ID = os.getenv("MS_CLIENT_ID")
MS_CLIENT_SECRET = os.getenv("MS_CLIENT_SECRET")
MS_DRIVE_ID = os.getenv("MS_DRIVE_ID")  # ID of the SharePoint/OneDrive drive to upload to

async def get_graph_token() -> str:
    """Acquires an app-only access token for Microsoft Graph."""
    if not all([MS_TENANT_ID, MS_CLIENT_ID, MS_CLIENT_SECRET]):
        raise HTTPException(status_code=500, detail="Credenziali Microsoft Graph non configurate.")

    url = f"https://login.microsoftonline.com/{MS_TENANT_ID}/oauth2/v2.0/token"
    data = {
        "client_id": MS_CLIENT_ID,
        "client_secret": MS_CLIENT_SECRET,
        "scope": "https://graph.microsoft.com/.default",
        "grant_type": "client_credentials"
    }
    
    async with httpx.AsyncClient() as client:
        response = await client.post(url, data=data)
        if response.status_code != 200:
            print(f"Error fetching token: {response.text}")
            raise HTTPException(status_code=500, detail="Errore di autenticazione con Microsoft Graph")
        return response.json()["access_token"]

async def upload_file_to_drive(file_content: bytes, file_name: str, folder_name: str) -> str:
    """
    Uploads a file to a specific folder in Microsoft Drive.
    folder_name can be the event name (e.g. "Assemblea Straordinaria").
    Returns the webUrl of the uploaded file.
    """
    if not MS_DRIVE_ID:
        raise HTTPException(status_code=500, detail="MS_DRIVE_ID non configurato.")

    token = await get_graph_token()
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/octet-stream"
    }

    # API Endpoint per caricare un file in una cartella specifica.
    # Usa l'upload per file < 4MB (put content). Per le deleghe (di solito 1 pagina PDF) va bene.
    # Costruiamo il percorso: /cartella/file.pdf
    # Sostituiamo spazi e caratteri non validi per sicurezza
    safe_folder = folder_name.replace("/", "-").replace("\\", "-")
    safe_filename = file_name.replace("/", "-").replace("\\", "-")
    
    path = f"{safe_folder}/{safe_filename}"
    url = f"https://graph.microsoft.com/v1.0/drives/{MS_DRIVE_ID}/root:/{path}:/content"

    async with httpx.AsyncClient() as client:
        response = await client.put(url, headers=headers, content=file_content)
        
        if response.status_code not in (200, 201):
            print(f"Error uploading file: {response.text}")
            raise HTTPException(status_code=500, detail="Errore durante l'upload del file su Microsoft Drive")
            
        data = response.json()
        return data.get("webUrl", "")

async def create_and_upload_empty_excel(folder_name: str, file_name: str = "Raccolta_Dati.xlsx") -> str:
    """
    Creates an empty Excel file with headers and uploads it to the specified folder.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Raccolta Dati"
    headers = ["Nome", "Email", "Modalità", "Delegato", "Intolleranze", "Registrato Il"]
    ws.append(headers)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True)
    
    output = io.BytesIO()
    wb.save(output)
    file_content = output.getvalue()
    
    return await upload_file_to_drive(file_content, file_name, folder_name)

async def append_to_excel_on_drive(folder_name: str, file_name: str, row_data: dict) -> bool:
    """
    Downloads an Excel file from OneDrive, appends a row, and uploads it back.
    """
    if not MS_DRIVE_ID:
        return False
        
    token = await get_graph_token()
    headers = {"Authorization": f"Bearer {token}"}
    
    safe_folder = folder_name.replace("/", "-").replace("\\", "-")
    safe_filename = file_name.replace("/", "-").replace("\\", "-")
    path = f"{safe_folder}/{safe_filename}"
    url = f"https://graph.microsoft.com/v1.0/drives/{MS_DRIVE_ID}/root:/{path}:/content"
    
    async with httpx.AsyncClient(follow_redirects=True) as client:
        # Download
        res_get = await client.get(url, headers=headers)
        if res_get.status_code != 200:
            print(f"Failed to download Excel file for appending: {res_get.text}")
            return False
            
        existing_content = res_get.content
        
        # Modify
        try:
            wb = openpyxl.load_workbook(filename=io.BytesIO(existing_content))
            ws = wb.active
            
            row = [
                row_data.get("nome", ""),
                row_data.get("email", ""),
                row_data.get("modalita", ""),
                row_data.get("delega_a", ""),
                row_data.get("intolleranze", ""),
                datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
            ]
            ws.append(row)
            
            output = io.BytesIO()
            wb.save(output)
            new_content = output.getvalue()
        except Exception as e:
            print(f"Error appending to Excel locally: {e}")
            return False
            
        # Upload
        headers["Content-Type"] = "application/octet-stream"
        res_put = await client.put(url, headers=headers, content=new_content)
        if res_put.status_code not in (200, 201):
            print(f"Failed to re-upload Excel file: {res_put.text}")
            return False
            
        return True
